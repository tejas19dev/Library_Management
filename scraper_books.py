"""
FreeComputerBooks.com — Full Scraper (200 books, all categories)
================================================================
Collects up to 200 books from all category types, scraping:
  Title, Author, Publisher, Permission, Paperback, eBook,
  Language, ISBN-10, ISBN-13, Cover Image, Description, Source URL

Run:
    pip install requests beautifulsoup4 openpyxl
    python scraper_books.py

Output: books_data_full.xlsx
"""

import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL   = "https://freecomputerbooks.com"
SITEMAP    = BASE_URL + "/sitemap.html"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
DELAY      = 1.0   # seconds between requests
MAX_BOOKS  = 200   # stop after this many books

NON_BOOK_PATTERNS = [
    "Category", "sitemap", "recentbooks", "top-",
    "books2024", "books2025", "books2026", "books2023",
    "books2022", "books2021", "books202",
    "about_books", "books_faq", "otherBooks",
    "miscellaneousBooks", "miscBooks",
    "searchEngines", "webTools", "mobile.html",
]

books_data    = []
visited_books = set()

# ── HTTP helper ───────────────────────────────────────────────────────────────
def get(url, timeout=15):
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"    ✗ {url}: {e}")
        return None

# ── Filter: is this href an individual book page? ─────────────────────────────
def is_book_page(href):
    if not href or not href.endswith(".html"):
        return False
    if any(p in href for p in NON_BOOK_PATTERNS):
        return False
    if href.endswith("Books.html") or href.endswith("Category.html"):
        return False
    return True

# ── STEP 1: get all category listing pages from sitemap ──────────────────────
def get_all_category_pages():
    print(f"  Fetching sitemap …")
    r = get(SITEMAP)
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    pages = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        full = urljoin(BASE_URL + "/", href)
        if (BASE_URL in full and href.endswith(".html")
                and not any(p in href for p in [
                    "about_books", "books_faq", "otherBooks",
                    "searchEngines", "webTools", "tradepub"
                ])):
            pages.add(full)
    return list(pages)

# ── STEP 2: collect book links from one category page ────────────────────────
def get_book_links(listing_url):
    r = get(listing_url)
    if not r:
        return []
    soup = BeautifulSoup(r.text, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not is_book_page(href):
            continue
        full = urljoin(BASE_URL + "/", href)
        if full not in visited_books and BASE_URL in full:
            visited_books.add(full)
            links.append(full)
    return links

# ── STEP 3: scrape all fields from one book page ─────────────────────────────
def extract_field(soup, label):
    """
    Find a <li> whose text starts with  **Label:** and return the value part.
    The HTML renders as:  <li><strong>Label:</strong> Value text</li>
    """
    for li in soup.find_all("li"):
        text = li.get_text(strip=True)
        # e.g. "Author(s): Ayman Alheraki"
        pattern = rf"^{re.escape(label)}\s*[:(]\s*(.+)"
        m = re.match(pattern, text, re.IGNORECASE)
        if m:
            return m.group(1).strip().rstrip(")")
    return "N/A"

def scrape_book(url):
    r = get(url)
    if not r:
        return
    soup = BeautifulSoup(r.text, "html.parser")

    # ── Title (h1) ────────────────────────────────────────────────────────
    title = "N/A"
    h1 = soup.find("h1")
    if h1:
        title = h1.get_text(strip=True)

    # ── All metadata fields from the <li> list ────────────────────────────
    author     = extract_field(soup, "Author(s)")
    if author == "N/A":
        author = extract_field(soup, "Author")

    publisher  = extract_field(soup, "Publisher")
    permission = extract_field(soup, "Permission")   # e.g. "Free", "CC BY"
    paperback  = extract_field(soup, "Paperback")
    ebook      = extract_field(soup, "eBook")
    language   = extract_field(soup, "Language")
    isbn10     = extract_field(soup, "ISBN-10")
    isbn13     = extract_field(soup, "ISBN-13")

    # ── Cover image (/covers/ path on the site) ───────────────────────────
    cover_img = "N/A"
    for img in soup.find_all("img"):
        src = img.get("src", "")
        if "/covers/" in src:
            cover_img = urljoin(BASE_URL + "/", src)
            break
    if cover_img == "N/A" and title != "N/A":
        cover_img = (
            "https://www.google.com/search?tbm=isch&q="
            + quote_plus(title + " book cover")
        )

    # ── Description (paragraph after "Book Description" heading) ─────────
    description = "N/A"
    desc_hdr = soup.find(
        lambda t: t.name in ["strong", "b", "h2", "h3", "p"]
                  and "book description" in t.get_text(strip=True).lower()
    )
    if desc_hdr:
        for sib in desc_hdr.find_next_siblings():
            text = sib.get_text(strip=True)
            if text and len(text) > 30:
                description = text[:1500]
                break
    if description == "N/A":
        best = ""
        for p in soup.find_all("p"):
            t = p.get_text(strip=True)
            if len(t) > 80 and len(t) > len(best):
                best = t
        if best:
            description = best[:1500]

    books_data.append({
        "Title":           title,
        "Author":          author,
        "Publisher":       publisher,
        "Permission":      permission,
        "Paperback":       paperback,
        "eBook":           ebook,
        "Language":        language,
        "ISBN-10":         isbn10,
        "ISBN-13":         isbn13,
        "Cover Image URL": cover_img,
        "Description":     description,
        "Source URL":      url,
    })
    print(f"  [{len(books_data):>3}/{MAX_BOOKS}] {title[:65]}")

# ── STEP 4: save to styled Excel ─────────────────────────────────────────────
def save_excel(path="books_data_full.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Books"

    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dat_font  = Font(name="Arial", size=10)
    dat_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt  = [
        PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF"),
        PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0"),
    ]

    COLS = [
        "#", "Title", "Author", "Publisher", "Permission",
        "Paperback", "eBook", "Language", "ISBN-10", "ISBN-13",
        "Cover Image URL", "Description", "Source URL"
    ]
    WIDTHS = [4, 50, 35, 30, 20, 12, 12, 12, 14, 16, 60, 90, 55]

    for c, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for r, d in enumerate(books_data, 2):
        fill = alt[(r - 2) % 2]
        values = [
            r - 1, d["Title"], d["Author"], d["Publisher"], d["Permission"],
            d["Paperback"], d["eBook"], d["Language"], d["ISBN-10"], d["ISBN-13"],
            d["Cover Image URL"], d["Description"], d["Source URL"]
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = dat_font; cell.fill = fill
            cell.alignment = dat_align; cell.border = brd
        ws.row_dimensions[r].height = 70

    wb.save(path)
    print(f"\n✅  Saved {len(books_data)} books  →  {path}")

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 65)
    print(f"  FreeComputerBooks — Scraper  (limit: {MAX_BOOKS} books)")
    print("=" * 65)

    # 1. Category pages
    print("\n[1/4] Collecting category pages from sitemap …")
    listing_pages = get_all_category_pages()
    print(f"      → {len(listing_pages)} category pages found\n")

    # 2. Book links (stop collecting once we have enough)
    print("[2/4] Scanning category pages for book links …")
    all_book_links = []
    for i, page in enumerate(listing_pages, 1):
        links = get_book_links(page)
        all_book_links.extend(links)
        print(f"  ({i:>3}/{len(listing_pages)})  +{len(links):>3} links  "
              f"| total: {len(all_book_links)}  — {page.split('/')[-1]}")
        time.sleep(DELAY)
        # We collect from ALL categories but scrape only MAX_BOOKS
        # so keep scanning until done (links are deduplicated already)

    # Trim to MAX_BOOKS
    all_book_links = all_book_links[:MAX_BOOKS]
    print(f"\n      → Scraping first {len(all_book_links)} unique books\n")

    # 3. Scrape each book
    print("[3/4] Scraping book detail pages …")
    for link in all_book_links:
        scrape_book(link)
        time.sleep(DELAY)
    print(f"\n      → {len(books_data)} books scraped\n")

    # 4. Excel
    print("[4/4] Writing Excel file …")
    save_excel("books_data_full.xlsx")

if __name__ == "__main__":
    main()
